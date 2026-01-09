"""
Export API Endpoints

Generate professional deliverables in multiple formats:
- PDF for Investment Memos, Executive Briefs
- Excel for Decision Matrix, Risk Register
"""

from fastapi import APIRouter, HTTPException, Response
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Dict, List, Optional, Any
from datetime import datetime
import io
import json

router = APIRouter(prefix="/api/exports", tags=["exports"])


# =============================================================================
# REQUEST MODELS
# =============================================================================

class InvestmentMemoExportRequest(BaseModel):
    """Request to export Investment Memo as PDF"""
    memo: Dict[str, Any]
    include_charts: bool = True
    format: str = "pdf"  # pdf or html


class DecisionMatrixExportRequest(BaseModel):
    """Request to export Decision Matrix as Excel"""
    matrix: Dict[str, Any]
    include_formatting: bool = True


class RiskRegisterExportRequest(BaseModel):
    """Request to export Risk Register"""
    register: Dict[str, Any]
    format: str = "excel"  # excel or pdf


class ExecutiveBriefExportRequest(BaseModel):
    """Request to export Executive Brief"""
    brief: Dict[str, Any]
    format: str = "pdf"


# =============================================================================
# PDF GENERATION
# =============================================================================

def generate_investment_memo_pdf(memo: Dict[str, Any]) -> bytes:
    """
    Generate professional Investment Memo PDF.

    Maps the deliverables data structure to HTML template.
    Uses browser Print-to-PDF for reliable cross-platform PDF generation.
    """
    # Extract data from deliverables structure
    company_name = memo.get('company_name', 'Company')
    recommendation = memo.get('recommendation', 'CONDITIONAL')
    recommendation_summary = memo.get('recommendation_summary', '')
    confidence = memo.get('confidence_level', 0.5)
    date = memo.get('date', datetime.now().strftime('%Y-%m-%d'))

    # Thesis data
    thesis = memo.get('thesis', {})
    headline = thesis.get('headline', '')
    rationale = thesis.get('rationale', [])
    upside = thesis.get('upside_case', '')
    downside = thesis.get('downside_case', '')

    # Deal terms
    deal = memo.get('deal', {})
    investment_amount = deal.get('investment_amount', 0)
    valuation = deal.get('pre_money_valuation', 0)
    ownership = deal.get('ownership_percentage', 0)
    instrument = deal.get('instrument', 'SAFE')

    # Metrics
    metrics = memo.get('metrics', {})

    # Factor assessments
    team = memo.get('team_assessment', {})
    market = memo.get('market_assessment', {})
    traction = memo.get('traction_assessment', {})
    competitive = memo.get('competitive_assessment', {})

    # Risks
    top_risks = memo.get('top_risks', [])

    # Return metrics
    expected_multiple = memo.get('expected_return_multiple', 0)
    expected_irr = memo.get('expected_irr', 0)

    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>Investment Memo - {company_name}</title>
        <style>
            @page {{ size: A4; margin: 0.75in; }}
            @media print {{ body {{ -webkit-print-color-adjust: exact; print-color-adjust: exact; }} }}
            body {{
                font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
                font-size: 11pt;
                line-height: 1.5;
                color: #1a1a1a;
                max-width: 800px;
                margin: 0 auto;
                padding: 20px;
            }}
            .header {{
                border-bottom: 3px solid #5E6AD2;
                padding-bottom: 20px;
                margin-bottom: 30px;
            }}
            .logo {{ font-size: 20px; font-weight: bold; color: #5E6AD2; letter-spacing: -0.5px; }}
            .title {{ font-size: 32px; font-weight: bold; margin-top: 10px; color: #1a1a1a; }}
            .subtitle {{ color: #666; font-size: 13px; margin-top: 5px; }}
            .section {{ margin-bottom: 28px; page-break-inside: avoid; }}
            .section-title {{
                font-size: 12px;
                font-weight: 700;
                text-transform: uppercase;
                letter-spacing: 1px;
                color: #5E6AD2;
                margin-bottom: 12px;
                border-bottom: 1px solid #e5e7eb;
                padding-bottom: 8px;
            }}
            .recommendation-box {{
                background: linear-gradient(135deg, #5E6AD2 0%, #A371F7 100%);
                color: white;
                padding: 24px;
                border-radius: 12px;
                margin: 24px 0;
            }}
            .rec-label {{ font-size: 11px; opacity: 0.9; text-transform: uppercase; letter-spacing: 1px; }}
            .rec-value {{ font-size: 28px; font-weight: bold; margin-top: 8px; }}
            .rec-summary {{ font-size: 13px; margin-top: 12px; opacity: 0.95; line-height: 1.6; }}
            .metric-grid {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 12px; margin: 16px 0; }}
            .metric-box {{
                background: #f8f9fc;
                padding: 16px;
                border-radius: 10px;
                border: 1px solid #e5e7eb;
                text-align: center;
            }}
            .metric-label {{ font-size: 10px; color: #666; text-transform: uppercase; letter-spacing: 0.5px; }}
            .metric-value {{ font-size: 20px; font-weight: bold; color: #1a1a1a; margin-top: 4px; }}
            .thesis-headline {{
                font-size: 18px;
                font-weight: 600;
                color: #1a1a1a;
                background: #f8f9fc;
                padding: 16px;
                border-radius: 8px;
                border-left: 4px solid #5E6AD2;
                margin-bottom: 16px;
            }}
            .rationale-list {{ padding-left: 20px; }}
            .rationale-list li {{ margin: 8px 0; color: #374151; }}
            .factor-grid {{ display: grid; grid-template-columns: repeat(2, 1fr); gap: 16px; }}
            .factor-card {{
                background: #f8f9fc;
                padding: 16px;
                border-radius: 10px;
                border: 1px solid #e5e7eb;
            }}
            .factor-header {{ display: flex; justify-content: space-between; align-items: center; margin-bottom: 8px; }}
            .factor-name {{ font-weight: 600; color: #374151; }}
            .factor-score {{ font-size: 18px; font-weight: bold; }}
            .factor-detail {{ font-size: 11px; color: #666; margin-top: 4px; }}
            .risk-item {{
                padding: 12px 16px;
                border-left: 4px solid #EF4444;
                background: #fef2f2;
                margin: 10px 0;
                border-radius: 0 8px 8px 0;
            }}
            .risk-title {{ font-weight: 600; color: #991b1b; font-size: 12px; }}
            .risk-mitigation {{ font-size: 11px; color: #666; margin-top: 4px; }}
            .deal-table {{ width: 100%; border-collapse: collapse; margin: 12px 0; }}
            .deal-table th, .deal-table td {{ padding: 12px; text-align: left; border-bottom: 1px solid #e5e7eb; }}
            .deal-table th {{ font-weight: 600; font-size: 11px; text-transform: uppercase; color: #666; background: #f8f9fc; }}
            .deal-table td {{ font-size: 13px; }}
            .footer {{
                margin-top: 40px;
                padding-top: 20px;
                border-top: 2px solid #e5e7eb;
                font-size: 10px;
                color: #666;
                text-align: center;
            }}
            .footer .tagline {{ font-style: italic; color: #5E6AD2; margin-top: 8px; }}
            .confidence-badge {{
                display: inline-block;
                padding: 4px 12px;
                border-radius: 20px;
                font-size: 11px;
                font-weight: 600;
                background: rgba(255,255,255,0.2);
                margin-top: 8px;
            }}
        </style>
    </head>
    <body>
        <div class="header">
            <div class="logo">POPULOUS</div>
            <div class="title">{company_name}</div>
            <div class="subtitle">Investment Memorandum | Prepared for Investment Committee | {date}</div>
        </div>

        <div class="recommendation-box">
            <div class="rec-label">Investment Recommendation</div>
            <div class="rec-value">{recommendation.replace('_', ' ')}</div>
            <div class="confidence-badge">{confidence*100:.0f}% Model Confidence</div>
            <div class="rec-summary">{recommendation_summary}</div>
        </div>

        <div class="section">
            <div class="section-title">Investment Thesis</div>
            <div class="thesis-headline">{headline}</div>
            <ul class="rationale-list">
                {''.join([f'<li>{r}</li>' for r in rationale[:4]])}
            </ul>
        </div>

        <div class="section">
            <div class="section-title">Key Metrics</div>
            <div class="metric-grid">
                <div class="metric-box">
                    <div class="metric-label">Unicorn Probability</div>
                    <div class="metric-value" style="color: #A371F7;">{metrics.get('unicorn_probability', 'N/A')}</div>
                </div>
                <div class="metric-box">
                    <div class="metric-label">Expected Valuation</div>
                    <div class="metric-value">{metrics.get('expected_valuation', 'N/A')}</div>
                </div>
                <div class="metric-box">
                    <div class="metric-label">Expected Return</div>
                    <div class="metric-value" style="color: #22C55E;">{expected_multiple:.1f}x</div>
                </div>
                <div class="metric-box">
                    <div class="metric-label">Expected IRR</div>
                    <div class="metric-value">{expected_irr*100:.0f}%</div>
                </div>
            </div>
        </div>

        <div class="section">
            <div class="section-title">Factor Analysis</div>
            <div class="factor-grid">
                <div class="factor-card">
                    <div class="factor-header">
                        <span class="factor-name">Team</span>
                        <span class="factor-score" style="color: #A371F7;">{team.get('score', 0)*100:.0f}/100</span>
                    </div>
                    <div class="factor-detail">Founders: {', '.join(team.get('founders', ['N/A'])[:2])}</div>
                    <div class="factor-detail">Strengths: {', '.join(team.get('strengths', ['N/A'])[:2])}</div>
                </div>
                <div class="factor-card">
                    <div class="factor-header">
                        <span class="factor-name">Market</span>
                        <span class="factor-score" style="color: #3FB950;">{market.get('score', 0)*100:.0f}/100</span>
                    </div>
                    <div class="factor-detail">TAM: {market.get('tam', 'N/A')}</div>
                    <div class="factor-detail">Growth: {market.get('growth_rate', 'N/A')}</div>
                </div>
                <div class="factor-card">
                    <div class="factor-header">
                        <span class="factor-name">Traction</span>
                        <span class="factor-score" style="color: #58A6FF;">{traction.get('score', 0)*100:.0f}/100</span>
                    </div>
                    <div class="factor-detail">{traction.get('funding', 'N/A')}</div>
                    <div class="factor-detail">{traction.get('stage', 'N/A')}</div>
                </div>
                <div class="factor-card">
                    <div class="factor-header">
                        <span class="factor-name">Competitive</span>
                        <span class="factor-score" style="color: #F59E0B;">Moat: {competitive.get('moat', 'N/A')}</span>
                    </div>
                    <div class="factor-detail">{competitive.get('differentiation', 'N/A')}</div>
                </div>
            </div>
        </div>

        <div class="section">
            <div class="section-title">Deal Terms</div>
            <table class="deal-table">
                <tr><th>Parameter</th><th>Value</th></tr>
                <tr><td>Investment Amount</td><td>${investment_amount/1e6:.2f}M</td></tr>
                <tr><td>Pre-Money Valuation</td><td>${valuation/1e6:.1f}M</td></tr>
                <tr><td>Ownership</td><td>{ownership:.1f}%</td></tr>
                <tr><td>Instrument</td><td>{instrument}</td></tr>
            </table>
        </div>

        <div class="section">
            <div class="section-title">Key Risks & Mitigations</div>
            {''.join([f'<div class="risk-item"><div class="risk-title">{r.get("risk", "Risk")}</div><div class="risk-mitigation">Mitigation: {r.get("mitigation", "N/A")}</div></div>' for r in top_risks[:3]])}
        </div>

        <div class="section">
            <div class="section-title">Scenario Analysis</div>
            <p><strong>Upside Case:</strong> {upside}</p>
            <p><strong>Downside Case:</strong> {downside}</p>
        </div>

        <div class="footer">
            <p>Generated by Populous Decision Intelligence Platform | CONFIDENTIAL</p>
            <p class="tagline">"There's a future where you win; we engineer that for you."</p>
        </div>
    </body>
    </html>
    """

    return html_content.encode('utf-8')


def generate_executive_brief_pdf(brief: Dict[str, Any]) -> bytes:
    """Generate Executive Brief PDF - one page summary for C-suite"""
    # Extract data from deliverables structure
    title = brief.get('title', 'Executive Brief')
    company_name = title.replace(' - Investment Decision Brief', '') if ' - ' in title else brief.get('company_name', 'Company')
    date = brief.get('date', datetime.now().strftime('%Y-%m-%d'))
    prepared_for = brief.get('prepared_for', 'Investment Committee')

    recommendation = brief.get('recommendation', 'CONDITIONAL')
    recommendation_text = brief.get('recommendation_text', '')
    confidence = brief.get('confidence', 0.5)
    deadline = brief.get('deadline', 'TBD')
    decision_required = brief.get('decision_required', '')

    key_points = brief.get('key_points', [])
    key_metrics = brief.get('key_metrics', {})
    top_risk = brief.get('top_risk', '')
    risk_mitigation = brief.get('risk_mitigation', '')

    alternatives = brief.get('alternatives', [])
    if_approved = brief.get('if_approved', [])
    if_rejected = brief.get('if_rejected', [])

    # Determine badge color
    badge_colors = {
        'STRONG_YES': ('#22C55E', 'white'),
        'YES': ('#86EFAC', '#166534'),
        'CONDITIONAL': ('#FCD34D', '#854D0E'),
        'NO': ('#EF4444', 'white'),
        'STRONG_NO': ('#991B1B', 'white')
    }
    bg_color, text_color = badge_colors.get(recommendation, ('#6B7280', 'white'))

    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>Executive Brief - {company_name}</title>
        <style>
            @page {{ size: A4; margin: 0.75in; }}
            @media print {{ body {{ -webkit-print-color-adjust: exact; print-color-adjust: exact; }} }}
            body {{
                font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
                font-size: 10pt;
                line-height: 1.5;
                color: #1a1a1a;
                max-width: 800px;
                margin: 0 auto;
                padding: 20px;
            }}
            .header {{
                display: flex;
                justify-content: space-between;
                align-items: flex-start;
                border-bottom: 3px solid #5E6AD2;
                padding-bottom: 16px;
                margin-bottom: 24px;
            }}
            .logo {{ font-size: 18px; font-weight: bold; color: #5E6AD2; }}
            .title {{ font-size: 24px; font-weight: bold; margin-top: 4px; }}
            .subtitle {{ color: #666; font-size: 11px; margin-top: 4px; }}
            .recommendation-badge {{
                display: inline-block;
                padding: 10px 20px;
                border-radius: 8px;
                font-weight: bold;
                font-size: 14px;
                background: {bg_color};
                color: {text_color};
            }}
            .decision-box {{
                background: #f8f9fc;
                border: 2px solid #5E6AD2;
                border-radius: 10px;
                padding: 16px;
                margin-bottom: 20px;
            }}
            .decision-label {{ font-size: 10px; text-transform: uppercase; color: #5E6AD2; font-weight: 600; }}
            .decision-text {{ font-size: 14px; font-weight: 600; margin-top: 4px; }}
            .metrics-grid {{ display: grid; grid-template-columns: repeat(5, 1fr); gap: 10px; margin: 16px 0; }}
            .metric {{
                padding: 12px;
                background: #f8f9fc;
                border-radius: 8px;
                text-align: center;
                border: 1px solid #e5e7eb;
            }}
            .metric-label {{ font-size: 9px; color: #666; text-transform: uppercase; letter-spacing: 0.5px; }}
            .metric-value {{ font-size: 16px; font-weight: bold; margin-top: 4px; color: #1a1a1a; }}
            .section {{ margin-bottom: 20px; }}
            .section-title {{
                font-size: 11px;
                font-weight: 700;
                text-transform: uppercase;
                letter-spacing: 1px;
                color: #5E6AD2;
                margin-bottom: 10px;
                padding-bottom: 6px;
                border-bottom: 1px solid #e5e7eb;
            }}
            .key-point {{
                padding: 10px 14px;
                background: #f8f9fc;
                border-left: 3px solid #5E6AD2;
                margin: 8px 0;
                font-size: 11px;
                border-radius: 0 6px 6px 0;
            }}
            .risk-box {{
                background: #fef2f2;
                border: 1px solid #fecaca;
                border-radius: 8px;
                padding: 14px;
            }}
            .risk-title {{ font-weight: 600; color: #991b1b; font-size: 12px; }}
            .risk-mitigation {{ font-size: 11px; color: #666; margin-top: 6px; }}
            .two-column {{ display: grid; grid-template-columns: 1fr 1fr; gap: 16px; }}
            .action-list {{ padding-left: 16px; }}
            .action-list li {{ font-size: 11px; margin: 6px 0; color: #374151; }}
            .footer {{
                margin-top: 30px;
                padding-top: 16px;
                border-top: 2px solid #e5e7eb;
                font-size: 9px;
                color: #666;
                text-align: center;
            }}
        </style>
    </head>
    <body>
        <div class="header">
            <div>
                <div class="logo">POPULOUS</div>
                <div class="title">{company_name}</div>
                <div class="subtitle">Executive Brief | {prepared_for} | {date}</div>
            </div>
            <div class="recommendation-badge">
                {recommendation.replace('_', ' ')}
            </div>
        </div>

        <div class="decision-box">
            <div class="decision-label">Decision Required</div>
            <div class="decision-text">{decision_required}</div>
            <div style="font-size: 11px; color: #666; margin-top: 8px;">Deadline: {deadline}</div>
        </div>

        <div class="section">
            <p style="font-size: 12px; line-height: 1.6;">{recommendation_text}</p>
            <div style="margin-top: 8px;">
                <span style="font-size: 11px; background: #5E6AD220; color: #5E6AD2; padding: 4px 10px; border-radius: 12px; font-weight: 600;">
                    {confidence*100:.0f}% Model Confidence
                </span>
            </div>
        </div>

        <div class="section">
            <div class="section-title">Key Metrics</div>
            <div class="metrics-grid">
                {''.join([f'<div class="metric"><div class="metric-label">{k}</div><div class="metric-value">{v}</div></div>' for k, v in list(key_metrics.items())[:5]])}
            </div>
        </div>

        <div class="section">
            <div class="section-title">Key Points</div>
            {''.join([f'<div class="key-point">{point}</div>' for point in key_points[:5]])}
        </div>

        <div class="section">
            <div class="section-title">Top Risk</div>
            <div class="risk-box">
                <div class="risk-title">{top_risk}</div>
                <div class="risk-mitigation">Mitigation: {risk_mitigation}</div>
            </div>
        </div>

        <div class="section">
            <div class="section-title">Next Steps</div>
            <div class="two-column">
                <div>
                    <strong style="font-size: 11px; color: #22C55E;">If Approved:</strong>
                    <ul class="action-list">
                        {''.join([f'<li>{step}</li>' for step in if_approved[:3]])}
                    </ul>
                </div>
                <div>
                    <strong style="font-size: 11px; color: #EF4444;">If Rejected:</strong>
                    <ul class="action-list">
                        {''.join([f'<li>{step}</li>' for step in if_rejected[:3]])}
                    </ul>
                </div>
            </div>
        </div>

        <div class="footer">
            <p>Generated by Populous Decision Intelligence Platform | CONFIDENTIAL</p>
        </div>
    </body>
    </html>
    """

    return html_content.encode('utf-8')


# =============================================================================
# EXCEL GENERATION
# =============================================================================

def generate_decision_matrix_excel(matrix: Dict[str, Any]) -> bytes:
    """Generate Decision Matrix Excel file"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Decision Matrix"

        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Title
        ws['A1'] = f"Decision Matrix - {matrix.get('batch_name', 'Batch')}"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:H1')

        ws['A2'] = f"Generated: {matrix.get('generated_at', datetime.now().strftime('%Y-%m-%d'))}"
        ws['A2'].font = Font(italic=True, color="666666")

        # Headers
        headers = ['Company', 'Probability', 'Team', 'Market', 'Traction', 'Timing', 'Capital', 'Recommendation']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        companies = matrix.get('companies', [])
        for row_idx, company in enumerate(companies, 5):
            ws.cell(row=row_idx, column=1, value=company.get('name', '')).border = border
            ws.cell(row=row_idx, column=2, value=f"{company.get('probability', 0)*100:.1f}%").border = border
            ws.cell(row=row_idx, column=3, value=f"{company.get('team', 0)*100:.0f}%").border = border
            ws.cell(row=row_idx, column=4, value=f"{company.get('market', 0)*100:.0f}%").border = border
            ws.cell(row=row_idx, column=5, value=f"{company.get('traction', 0)*100:.0f}%").border = border
            ws.cell(row=row_idx, column=6, value=f"{company.get('timing', 0)*100:.0f}%").border = border
            ws.cell(row=row_idx, column=7, value=f"{company.get('capital', 0)*100:.0f}%").border = border

            rec = company.get('recommendation', 'PASS')
            rec_cell = ws.cell(row=row_idx, column=8, value=rec)
            rec_cell.border = border
            if rec == 'STRONG_YES':
                rec_cell.fill = PatternFill(start_color="22C55E", end_color="22C55E", fill_type="solid")
                rec_cell.font = Font(color="FFFFFF", bold=True)
            elif rec == 'YES':
                rec_cell.fill = PatternFill(start_color="86EFAC", end_color="86EFAC", fill_type="solid")
            elif rec == 'CONDITIONAL':
                rec_cell.fill = PatternFill(start_color="FCD34D", end_color="FCD34D", fill_type="solid")

        # Auto-width columns
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 15
        ws.column_dimensions['A'].width = 25

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    except ImportError:
        # Return CSV as fallback
        csv_content = "Company,Probability,Team,Market,Traction,Timing,Capital,Recommendation\n"
        for company in matrix.get('companies', []):
            csv_content += f"{company.get('name', '')},{company.get('probability', 0)*100:.1f}%,"
            csv_content += f"{company.get('team', 0)*100:.0f}%,{company.get('market', 0)*100:.0f}%,"
            csv_content += f"{company.get('traction', 0)*100:.0f}%,{company.get('timing', 0)*100:.0f}%,"
            csv_content += f"{company.get('capital', 0)*100:.0f}%,{company.get('recommendation', 'PASS')}\n"
        return csv_content.encode('utf-8')


def generate_risk_register_excel(register: Dict[str, Any]) -> bytes:
    """Generate Risk Register Excel file"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Risk Register"

        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Title
        ws['A1'] = f"Risk Register - {register.get('company_name', 'Company')}"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:F1')

        ws['A2'] = f"Generated: {register.get('generated_at', datetime.now().strftime('%Y-%m-%d'))}"
        ws['A2'].font = Font(italic=True, color="666666")

        # Headers
        headers = ['Risk Category', 'Description', 'Probability', 'Impact', 'Severity', 'Mitigation']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        # Risk data
        risks = register.get('risks', [])
        for row_idx, risk in enumerate(risks, 5):
            ws.cell(row=row_idx, column=1, value=risk.get('category', '')).border = border
            ws.cell(row=row_idx, column=2, value=risk.get('description', '')).border = border
            ws.cell(row=row_idx, column=3, value=risk.get('probability', '')).border = border
            ws.cell(row=row_idx, column=4, value=risk.get('impact', '')).border = border

            severity_cell = ws.cell(row=row_idx, column=5, value=risk.get('severity', ''))
            severity_cell.border = border
            severity = risk.get('severity', '')
            if severity == 'CRITICAL':
                severity_cell.fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
                severity_cell.font = Font(color="FFFFFF", bold=True)
            elif severity == 'HIGH':
                severity_cell.fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
            elif severity == 'MEDIUM':
                severity_cell.fill = PatternFill(start_color="FCD34D", end_color="FCD34D", fill_type="solid")

            ws.cell(row=row_idx, column=6, value=risk.get('mitigation', '')).border = border

        # Auto-width
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 40

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    except ImportError:
        csv_content = "Category,Description,Probability,Impact,Severity,Mitigation\n"
        for risk in register.get('risks', []):
            csv_content += f'"{risk.get("category", "")}","{risk.get("description", "")}",'
            csv_content += f'"{risk.get("probability", "")}","{risk.get("impact", "")}",'
            csv_content += f'"{risk.get("severity", "")}","{risk.get("mitigation", "")}"\n'
        return csv_content.encode('utf-8')


# =============================================================================
# ENDPOINTS
# =============================================================================

@router.post("/investment-memo/pdf")
async def export_investment_memo_pdf(request: InvestmentMemoExportRequest):
    """Export Investment Memo as PDF-ready HTML"""
    try:
        # Generate HTML that can be printed to PDF by browser
        html_bytes = generate_investment_memo_pdf(request.memo)

        return Response(
            content=html_bytes,
            media_type="text/html",
            headers={
                "Content-Disposition": f"inline; filename=investment_memo_{request.memo.get('company_name', 'company').replace(' ', '_')}.html"
            }
        )
    except Exception as e:
        raise HTTPException(500, f"Failed to generate document: {str(e)}")


@router.post("/executive-brief/pdf")
async def export_executive_brief_pdf(request: ExecutiveBriefExportRequest):
    """Export Executive Brief as PDF-ready HTML"""
    try:
        html_bytes = generate_executive_brief_pdf(request.brief)

        return Response(
            content=html_bytes,
            media_type="text/html",
            headers={
                "Content-Disposition": f"inline; filename=executive_brief_{request.brief.get('company_name', 'company').replace(' ', '_')}.html"
            }
        )
    except Exception as e:
        raise HTTPException(500, f"Failed to generate document: {str(e)}")


@router.post("/decision-matrix/excel")
async def export_decision_matrix_excel(request: DecisionMatrixExportRequest):
    """Export Decision Matrix as Excel"""
    try:
        excel_bytes = generate_decision_matrix_excel(request.matrix)

        # Check if we got Excel or CSV fallback
        is_excel = excel_bytes[:4] == b'PK\x03\x04'  # XLSX magic bytes

        return Response(
            content=excel_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if is_excel else "text/csv",
            headers={
                "Content-Disposition": f"attachment; filename=decision_matrix_{request.matrix.get('batch_name', 'batch').replace(' ', '_')}.{'xlsx' if is_excel else 'csv'}"
            }
        )
    except Exception as e:
        raise HTTPException(500, f"Failed to generate Excel: {str(e)}")


@router.post("/risk-register/excel")
async def export_risk_register_excel(request: RiskRegisterExportRequest):
    """Export Risk Register as Excel"""
    try:
        if request.format == "excel":
            excel_bytes = generate_risk_register_excel(request.register)
            is_excel = excel_bytes[:4] == b'PK\x03\x04'

            return Response(
                content=excel_bytes,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if is_excel else "text/csv",
                headers={
                    "Content-Disposition": f"attachment; filename=risk_register_{request.register.get('company_name', 'company').replace(' ', '_')}.{'xlsx' if is_excel else 'csv'}"
                }
            )
        else:
            # PDF format not implemented yet for risk register
            raise HTTPException(400, "PDF format not yet supported for Risk Register")
    except Exception as e:
        raise HTTPException(500, f"Failed to generate export: {str(e)}")
